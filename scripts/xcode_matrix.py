#!/usr/bin/env python3
"""Three-code parameter matrix: FLASH vs PSC vs WarpX, with what is preserved and what is not.

    /opt/anaconda3/envs/physics/bin/python scripts/xcode_matrix.py

Writes media/xcode/xcode_matrix.{csv,xlsx,html}.

WHY THIS EXISTS. The Phase-4 campaign spent a dozen runs attributing a 3x plume-temperature
gap to physics, when the gap is a PARAMETER difference the normalised axes hid. Each row here
carries its own provenance so a claim can be checked against the file it came from rather than
against memory. STATUS is the point of the sheet:

  SAME    identical across all three by design -- if one of these ever differs, that is a bug
  PIC     the two PIC codes agree, FLASH differs because it is a fluid code (expected)
  DIFFERS the codes disagree in a way that MATTERS and is not merely a unit convention
  N/A     the quantity has no meaning for that code

WHICH PSC RUN THE COLUMN IS  (changed 2026-08-27, and the reason this file was revised).
PSC has TWO independent normalisation knobs and the second one is easy to miss:

    ReducedMassRatio = 100                 sets d_i0 and the mass unit
    ReducedSoL       = 3000 eV / m_e c^2   sets the temperature unit, the clock and collisions

There are two PSC legs of the same problem. `run_ourflash` runs the paper's ReducedSoL = 0.05,
i.e. m_e c^2 = 60 keV -- an EXTRA, non-similar reduction of c by sqrt(511/60) = 2.92 on top of
the mass reduction, which also leaves it 72.5x over-collisional. `run_ourflash_511keV` runs
3000/511000, i.e. the REAL electron rest energy, which for an 18.36x electron means c/4.285
exactly -- the similarity-consistent choice -- and a collision rate of exactly 1.00x physical.
K_length and K_mass are IDENTICAL between the two legs, so every length and every physical mass
below is common to both; what moves is the temperature unit (8.52x), the clock (2.92x) and the
collisionality (72.5x).

THE PSC COLUMN IS NOW `run_ourflash_511keV` THROUGHOUT. The previous version of this sheet was
not one run: `plume T_e` = 509 eV came from the 511 keV leg while `f_abs` = 0.764 and
`duration analysed` = 166 ps came from the 60 keV leg 66 ps later -- and `T_ss`, hence the
`T_e/T_ss` = 0.74 that the sheet turned on, was computed from that cross-run pairing. Worse,
0.764 was the last finite `fabs` the 60 keV run printed: one step after its only NaN, at the
glibc abort in `pic_move_part_z_` that killed it at 18% of its planned length.

Two caveats the reader is owed, both recorded in the rows themselves:
  * the legs are not tau-matched. PSC reaches 2.69 tau_own, WarpX 5.39, FLASH 26.96, and PSC's
    plume T_e is still rising at cutoff -- so 509 eV is a lower bound.
  * PSC's f_abs is a spiky instantaneous diagnostic (0.33 -> 1.00 -> 0.48 across the run) and
    its 511 keV plume T_e is non-monotonic (459 -> 279 -> 509 eV), the dip coinciding with
    f_abs pinned at 1.000. The endpoint sits on a recovery limb.
At the MATCHED time of 99.7 ps the two PSC legs agree to 5% on T_e/T_ss (0.97 at 60 keV, 1.02
at 511 keV), which is the sense in which the switch is a correction of bookkeeping and not of
physics -- it just removes the paper's extra c reduction from the WarpX comparison.

The single DIFFERS row that explains the campaign is `m_ion (physical)`: PSC reaches
m_i/m_e = 100 by making the ELECTRON 18.4x heavier and keeps a REAL aluminium ion, while WarpX
makes the ION 18.2x lighter and keeps a real electron. Same ratio, opposite implementation,
and only the ion mass sets C_S and the Manheimer steady state.
"""
import csv, os

OUT = "media/xcode"
PSC_LEG = "run_ourflash_511keV"        # ReducedSoL = 3000/511000; see the module docstring

ROWS = [
    # (section, parameter, FLASH, PSC, WarpX, status, provenance)
    ("Laser", "wavelength", "1.064 um", "1.064 um", "1.064 um", "SAME", "flash.par ed_wavelength_1; INIT_param.f lw; config laser.wavelength_um"),
    ("Laser", "intensity", "1e13 W/cm^2", "1e13 W/cm^2", "1e13 W/cm^2", "SAME", "flash.par ed_power 1e20 erg/s on unit cross-section; BOTH PSC legs log I0 = 1.00000002E+20 erg/s/cm^2; config 1.0e17 W/m^2"),
    ("Laser", "n_cr", "9.849e26 m^-3", "9.849e26 m^-3", "9.849e26 m^-3", "SAME", "PSC K_density = 9.8489880e26, identical in both legs; deck ncr = epsilon0*m_e*omega0^2/q_e^2"),
    ("Laser", "incidence / profile", "normal, uniform", "normal, uniform", "normal, uniform", "SAME", "flash.par crossSectionFunctionType=uniform; config beam.profile=uniform"),
    ("Laser", "absorption model", "ray-trace IB (EnergyDeposition)", "ray-trace IB (Hyder module)", "ray-trace IB (LaserDeposition)", "SAME", "same KIND of operator in all three"),
    ("Laser", "Z_eff", "13", "13", "13", "SAME", "INIT_param.f Z_eff=ZZ1=13; config laser.Z_eff"),
    ("Laser", "lnLambda, LASER operator", "not extracted", "NRL per-cell, floor 1", "NRL per-cell, floor 1", "PIC", "PSC get_lnlambda (PIC_part_heating.F90:882); LaserDeposition.cpp coulombLog mode 1 -- SAME formula and floor. NOT the collision operator's value, see Plasma"),
    ("Laser", "lnLambda in plume (measured)", "n/a", "6.27 (336 cells, 4.95-7.17)", "4.75 (168 cells, max 6.03)", "DIFFERS", "NRL recomputed per-cell from each code's own dumped n_e and T_e -- bit-exact with what its laser operator saw. PSC at 99.66 ps; 60 keV leg gives 6.25, so this is NOT a leg difference"),
    ("Laser", "pulse shape", "0.1 ns rise, flat to 1 ns", "constant (starts at handoff)", "constant (starts at handoff)", "PIC", "PIC runs start from the 0.1 ns FLASH state, so they see only the flat top"),

    ("Mass", "m_e (physical)", "real", "18.36x real", "real", "DIFFERS", "K_mass = hydr_mass_phys/ReducedMassRatio = 1.6726e-29 kg -- has NO ReducedSoL, so IDENTICAL in both PSC legs"),
    ("Mass", "m_e c^2", "n/a (fluid)", "511 keV", "511 keV", "PIC", "run_ourflash_511keV logs K_temperature = 511000.000, the REAL electron rest energy. The paper's leg runs 60 keV -- this is the row the column switch was about"),
    ("Mass", "c (speed of light)", "n/a (fluid)", "0.2333 c", "1.000 c", "DIFFERS", "c = K_length/K_time = 7.261e-7/1.0381e-14 = 6.994e7 m/s = c/4.285 = c/sqrt(18.36) -- EXACTLY the value that makes m_e c^2 real for an 18.36x electron. The 60 keV leg runs 0.0799 c, a further 2.92x cut that is NOT part of the similarity transform"),
    ("Mass", "m_ion (physical)  <<< THE ONE", "4.480e-26 kg (real Al)", "4.513e-26 kg (1.008x real Al)", "2.458e-27 kg (0.055x real Al)", "DIFFERS", "PSC MMi1*K_mass = 26.9815*hydr_mass_phys, ReducedMassRatio CANCELS, ReducedSoL absent -- same in both legs; WarpX Mi = mass_ratio*m_e"),
    ("Mass", "m_i / m_e", "49542", "2698", "2698", "PIC", "the two PIC codes DO agree here -- which is why the ion-mass difference was invisible"),
    ("Mass", "reduction implemented by", "n/a", "heavier ELECTRON, c cut 4.285x to match", "lighter ION, real c", "DIFFERS", "opposite directions, same m_i/m_e; PSC pays for its electron in the speed of light, WarpX in the ion mass"),

    ("Length", "d_e at n_cr", "n/a (fluid)", "0.7256 um", "0.1693 um (= lambda/2pi)", "DIFFERS", "d_e ~ sqrt(m_e); PSC K_length = 7.26098e-7 m, identical in both legs (K_length = DI0_phys/sqrt(ReducedMassRatio), no ReducedSoL)"),
    ("Length", "d_i0 (proton skin depth)", "7.256 um", "7.261 um", "1.693 um", "DIFFERS", "PSC DI0_phys uses the REAL proton; WarpX uses the reduced one"),
    ("Length", "initial target thickness", "50 um solid slab", "32.65 um (45 code-d_e)", "7.62 um (45 real-d_e)", "DIFFERS", "flash.par sim_targetHeight=5.0e-3 cm; both PIC decks say '45 d_e' and mean different lengths"),
    ("Length", "cell size", "0.781 um (AMR lrefine 4)", "0.145 um (0.2 code-d_e)", "0.0847 um (0.5 real-d_e)", "DIFFERS", "DELIVERY.md dx_min; PSC DZ=0.2; config dz_over_de=0.5"),
    ("Length", "domain", "800 um", "726 um (1000 code-d_e)", "423 um (2500 real-d_e)", "DIFFERS", "flash.par xmax=0.08 cm; PSC 5000 cells x 0.2 d_e; config axis lo/hi"),

    ("Time", "real time per unit tau", "37.098 ps", "37.098 ps (maps 1:1)", "2.028 ps", "DIFFERS", "tau = d_i0/C_S0. PSC TD0_phys is built from the REAL proton and has no ReducedSoL, so it is the same in both legs; WarpX tau uses the reduced ion"),
    ("Time", "timestep", "n/a (AMR/hydro)", "1.557 fs", "0.0989 fs", "DIFFERS", "PSC dt = 0.15*K_time from its run log; the 60 keV leg runs 4.544 fs. WarpX run.log 'Level 0: dt = 9.885055348e-17' at cfl 0.35 -- 15.7x smaller, the price of real c"),
    ("Time", "duration analysed", "1.0 ns (26.96 tau)", "99.7 ps (2.69 tau)", "10.93 ps (5.39 tau)", "DIFFERS", "PSC ran 64000 steps to t_FLASH 0.1996 ns and finished CLEAN. NOT tau-matched to WarpX, and PSC's plume T_e is still rising at cutoff -- its 509 eV is a lower bound"),

    ("Plasma", "peak density n_max", "795 n_cr (solid Al)", "10 n_cr", "10 n_cr", "PIC", "DELIVERY.md 795.5 n_cr exact; paper Appendix A caps PIC at 10"),
    ("Plasma", "initial condition", "cold 290 K slab", "FLASH at 0.1 ns", "FLASH at 0.1 ns", "PIC", "both PIC runs are handed off from the same FLASH snapshot; PSC's were regenerated into ic_ourflash_511keV because K_temperature IS m_e c^2"),
    ("Plasma", "IC mapped through", "n/a", "zeta -> z_code/10", "zeta -> z/DI0_W", "DIFFERS", "same zeta profile, physical extents differing by 4.285x"),
    ("Plasma", "chamber gas", "1e-10 g/cm^3", "none", "none", "PIC", "FLASH needs a floor density; PIC does not"),
    ("Plasma", "collision operator", "fluid transport", "binary Coulomb", "binary Coulomb (Perez)", "PIC", "config collisions.type"),
    ("Plasma", "collision lnLambda", "n/a", "8.28 global", "6.3 global", "DIFFERS", "CORRECTED: PSC's collision operator does NOT use the per-cell NRL. INIT_param.f:584 sets lnlambda = 23 - ln(sqrt(n_cr)*Z_eff/T^1.5) ONCE at n_cr and 3000 eV, floored at 1, and it enters only through nudt0. Both codes are global constants; 8.28 vs config collisions.coulomb_log 6.3"),
    ("Plasma", "collision rate vs physical", "n/a", "1.00x", "1.00x", "PIC", "INIT_param.f:594 nudt0 *= ReducedSoL^2*(511000/temp_phys)^2 = (511000/K_temperature)^2, which is 1.0000 at 511 keV. The paper's 60 keV leg is 72.5x OVER-collisional (logged nudt0 1.4970e-6 vs 2.0639e-8) -- a second reason the 511 keV leg is the right comparison"),
    ("Plasma", "macroparticles", "n/a (fluid)", "~1e5 ppc at n_cr, weight-proportional", "500 ppc fixed weight", "DIFFERS", "different loading CONVENTION -- compare resolved dynamic range, not the number. PNUM TOT 2.793e6 in both PSC legs"),

    ("Result", "f_abs (measured)", "0.870", "0.475", "0.350", "DIFFERS", "final instantaneous fabs from the LASERDEP lines -- the SAME convention as WarpX's 0.350 (f_end, not the time-integrated mean, which is 0.583 for PSC). The old 0.764 was the 60 keV leg at 165.9 ps, one step after its only NaN and at the glibc abort that killed it"),
    ("Result", "plume T_e (measured)", "647 eV", "509 eV", "157.7 eV", "DIFFERS", "n-weighted over 0.05 < n_e/n_cr < 1.0, each leg on its own zeta. PSC 508.8 eV at 99.66 ps, 336 cells; the 60 keV leg gives 516.1 eV at the same time -- 1.4% apart, so the reduced c is NOT the discriminator"),
    ("Result", "T_ss for its OWN ion mass", "750 eV", "501 eV", "155 eV", "DIFFERS", "823 * (m_i/m_i,real)^(1/3) * f_abs^(2/3), with each leg's OWN f_abs on the row above. CAVEAT (2026-08-27): the f_abs^(2/3) term does NOT survive a mu-sweep -- reducing mr25/mr100/mr400 by it gives 1529/838/495 eV, a 3.1x spread where a valid reduction gives one constant, while mu^(1/3) alone fits to 2.3%. It behaves within one mass ratio (1-6%). So this row is soft ACROSS the three codes; the hard result is the matched-f_abs leg P4_lez_kin_clmatch, which agrees with PSC to 4.8%"),
    ("Result", "T_e / T_ss  <<< THE POINT", "0.86", "1.02", "1.02", "SAME", "ALL THREE sit at 0.86-1.02 of their own steady state and the two PIC codes land on the SAME number, so no code is anomalous -- but see the T_ss caveat above: this ratio inherits f_abs^(2/3). The claim does not depend on it. P4_lez_kin_clmatch matches PSC's f_abs experimentally (0.5629 vs 0.5833) and gives PSC/WarpX = 2.578 against mu^(1/3)'s 2.645, 4.8% apart on a 13.5% floor -- mu^(1/3) is the whole difference"),
]

HDR = ["Section", "Parameter", "FLASH", "PSC", "WarpX", "Status", "Provenance"]

def write_csv(path):
    with open(path, "w", newline="") as f:
        w = csv.writer(f); w.writerow(HDR); w.writerows(ROWS)

def write_xlsx(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    FILL = {"SAME":    PatternFill("solid", fgColor="C6EFCE"),
            "PIC":     PatternFill("solid", fgColor="FFF2CC"),
            "DIFFERS": PatternFill("solid", fgColor="F8CBAD"),
            "N/A":     PatternFill("solid", fgColor="E7E6E6")}
    FONT = {"SAME": Font(color="006100"), "PIC": Font(color="7F6000"),
            "DIFFERS": Font(color="9C0006", bold=True), "N/A": Font(color="808080")}
    wb = Workbook(); ws = wb.active; ws.title = "Three-code matrix"
    hf = PatternFill("solid", fgColor="1F3864")
    thin = Side(style="thin", color="BFBFBF"); bd = Border(bottom=thin)
    ws.append(HDR)
    for c in ws[1]:
        c.fill = hf; c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = Alignment(vertical="center", horizontal="center")
    ws.freeze_panes = "C2"
    last = None
    for r in ROWS:
        ws.append(list(r))
        i = ws.max_row
        st = r[5]
        for c in ws[i]:
            c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd
        ws.cell(i, 6).fill = FILL[st]; ws.cell(i, 6).font = FONT[st]
        ws.cell(i, 6).alignment = Alignment(horizontal="center", vertical="center")
        if st == "DIFFERS":
            for col in (3, 4, 5):
                ws.cell(i, col).fill = FILL["DIFFERS"]
        elif st == "SAME":
            for col in (3, 4, 5):
                ws.cell(i, col).fill = FILL["SAME"]
        ws.cell(i, 2).font = Font(bold=True)
        if r[0] != last:
            ws.cell(i, 1).font = Font(bold=True, color="1F3864"); last = r[0]
        else:
            ws.cell(i, 1).value = ""
    for col, wd in zip("ABCDEFG", (11, 30, 26, 30, 30, 10, 62)):
        ws.column_dimensions[col].width = wd
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 30
    # legend
    ws2 = wb.create_sheet("Legend")
    ws2.append(["Status", "Meaning"])
    for c in ws2[1]:
        c.fill = hf; c.font = Font(color="FFFFFF", bold=True)
    for k, v in (("SAME", "Identical across all three by design. If one of these ever differs, that is a bug."),
                 ("PIC", "The two PIC codes agree; FLASH differs because it is a fluid code. Expected."),
                 ("DIFFERS", "The codes disagree in a way that MATTERS and is not merely a unit convention."),
                 ("N/A", "The quantity has no meaning for that code.")):
        ws2.append([k, v]); ws2.cell(ws2.max_row, 1).fill = FILL[k]
        ws2.cell(ws2.max_row, 1).font = FONT[k]
        ws2.cell(ws2.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 12; ws2.column_dimensions["B"].width = 95
    for i in range(2, ws2.max_row + 1): ws2.row_dimensions[i].height = 30
    ws2.append([])
    ws2.append(["THE HEADLINE", "PSC reaches m_i/m_e = 100 by making the ELECTRON 18.4x heavier, keeping a REAL "
                "aluminium ion. WarpX makes the ION 18.2x lighter, keeping a real electron. Same ratio, opposite "
                "implementation. Only the ION mass sets C_S and the Manheimer steady state, so the whole 3x plume "
                "temperature gap follows -- and once each leg is measured against ITS OWN steady state, all three "
                "sit at 0.86-1.02 and no code is anomalous. The two PIC codes land on the SAME 1.02."])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True, color="9C0006")
    ws2.cell(ws2.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[ws2.max_row].height = 90
    ws2.append([])
    ws2.append(["THE PSC LEG", f"The PSC column is {PSC_LEG}, ReducedSoL = 3000/511000, i.e. m_e c^2 = 511 keV "
                "-- the REAL electron rest energy, c = c/4.285 exactly (the similarity-consistent value for an "
                "18.36x electron), and a collision rate of exactly 1.00x physical. The paper's leg, run_ourflash, "
                "runs 60 keV: an EXTRA 2.92x cut in c that is not part of the similarity transform, and 72.5x "
                "over-collisional. K_length and K_mass are identical between the legs, so every length and every "
                "physical mass in this sheet is common to both. An earlier version of this sheet was not one run "
                "-- T_e came from the 511 keV leg and f_abs from the 60 keV leg 66 ps later -- which is where the "
                "old T_e/T_ss = 0.74 came from."])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True, color="1F3864")
    ws2.cell(ws2.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[ws2.max_row].height = 135
    wb.save(path)

if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)
    write_csv(f"{OUT}/xcode_matrix.csv"); write_xlsx(f"{OUT}/xcode_matrix.xlsx")
    n = {s: sum(1 for r in ROWS if r[5] == s) for s in ("SAME", "PIC", "DIFFERS", "N/A")}
    print(f"wrote {OUT}/xcode_matrix.csv and .xlsx  ({len(ROWS)} rows)")
    print(f"  SAME {n['SAME']}   PIC {n['PIC']}   DIFFERS {n['DIFFERS']}")

# ---------------------------------------------------------------------------------------
# HTML view, generated from the SAME ROWS so the page and the spreadsheet cannot diverge.
# ---------------------------------------------------------------------------------------
HEAD = {"FLASH": ("radiation-hydrodynamics", "real Al ion", "0.86"),
        "PSC":   ("PIC, heavier electron + c/4.29", "real Al ion", "1.02"),
        "WarpX": ("PIC, lighter ion, real c", "0.055x Al ion", "1.02")}

def write_html(path):
    import html as H
    def cell(v):
        return H.escape(v).replace("^-3", "&#8315;&#179;").replace("^3", "&#179;") \
                          .replace("^2", "&#178;").replace("->", "&rarr;").replace("~", "&#8776;")
    body, last = [], None
    for sec, par, f, p, w, st, src in ROWS:
        flag = ""
        if "<<<" in par:
            par, tag = par.split("<<<"); flag = f'<span class="flag">{H.escape(tag.strip())}</span>'
        if sec != last:
            body.append(f'<tr class="sec"><th colspan="6" scope="rowgroup">{H.escape(sec)}</th></tr>')
            last = sec
        body.append(
            f'<tr class="s-{st.lower().replace("/","")}">'
            f'<th scope="row"><span class="par">{H.escape(par.strip())}</span>{flag}'
            f'<span class="src">{cell(src)}</span></th>'
            f'<td>{cell(f)}</td><td>{cell(p)}</td><td>{cell(w)}</td>'
            f'<td class="st"><span class="pill">{H.escape(st)}</span></td></tr>')
    chips = "".join(
        f'<div class="chip"><div class="ck">{k}</div><div class="cd">{d}</div>'
        f'<div class="cm">{m}</div><div class="cn">{r}</div>'
        f'<div class="cl">T<sub>e</sub> / T<sub>ss</sub></div></div>'
        for k, (d, m, r) in HEAD.items())
    open(path, "w").write(TEMPLATE.replace("@@CHIPS@@", chips).replace("@@ROWS@@", "\n".join(body)))

TEMPLATE = r"""<title>Three-Code Parameter Matrix</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Spectral:ital,wght@0,400;0,600;1,400&family=IBM+Plex+Sans:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap">
<style>
:root{
  --ground:#F6F8F9; --surface:#FFFFFF; --sunk:#EDF1F3;
  --ink:#14202A; --ink-2:#42565F; --ink-3:#7A8C95;
  --rule:#DAE3E7; --rule-2:#C3D2D8;
  --accent:#0E6E78; --accent-soft:#E2F0F1;
  --ok:#1B6E45; --ok-bg:#E4F1E9; --warn:#8A5B00; --warn-bg:#F8EEDC;
  --diff:#A32E22; --diff-bg:#F9E7E4;
  --shadow:0 1px 2px rgba(20,32,42,.05),0 8px 24px -12px rgba(20,32,42,.18);
}
@media (prefers-color-scheme:dark){:root:not([data-theme="light"]){
  --ground:#0C1216; --surface:#131C22; --sunk:#0F171C;
  --ink:#E4EBEF; --ink-2:#A2B4BC; --ink-3:#6E828B;
  --rule:#23313A; --rule-2:#2E4049;
  --accent:#4CBCC4; --accent-soft:#103336;
  --ok:#6FCB97; --ok-bg:#11291D; --warn:#E0B265; --warn-bg:#2A2113;
  --diff:#F0897B; --diff-bg:#2E1815;
  --shadow:0 1px 2px rgba(0,0,0,.4),0 8px 24px -12px rgba(0,0,0,.6);
}}
:root[data-theme="dark"]{
  --ground:#0C1216; --surface:#131C22; --sunk:#0F171C;
  --ink:#E4EBEF; --ink-2:#A2B4BC; --ink-3:#6E828B;
  --rule:#23313A; --rule-2:#2E4049;
  --accent:#4CBCC4; --accent-soft:#103336;
  --ok:#6FCB97; --ok-bg:#11291D; --warn:#E0B265; --warn-bg:#2A2113;
  --diff:#F0897B; --diff-bg:#2E1815;
  --shadow:0 1px 2px rgba(0,0,0,.4),0 8px 24px -12px rgba(0,0,0,.6);
}
*{box-sizing:border-box}
body{margin:0;background:var(--ground);color:var(--ink);
  font-family:"IBM Plex Sans",system-ui,-apple-system,sans-serif;
  font-size:15px;line-height:1.55;-webkit-font-smoothing:antialiased}
.wrap{max-width:1180px;margin:0 auto;padding:48px 24px 96px;display:flex;flex-direction:column;gap:38px}
.eyebrow{font-size:11.5px;letter-spacing:.14em;text-transform:uppercase;color:var(--accent);
  font-weight:600;margin:0 0 10px}
h1{font-family:Spectral,Georgia,serif;font-weight:600;font-size:clamp(30px,4.6vw,46px);
  line-height:1.1;margin:0 0 14px;text-wrap:balance;letter-spacing:-.012em}
.lede{font-family:Spectral,Georgia,serif;font-size:clamp(16px,2vw,19px);line-height:1.6;
  color:var(--ink-2);max-width:66ch;margin:0}
.lede em{color:var(--ink);font-style:italic}
.chips{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px}
.chip{background:var(--surface);border:1px solid var(--rule);border-radius:3px;padding:16px 18px;
  box-shadow:var(--shadow);display:flex;flex-direction:column;gap:2px}
.ck{font-family:"IBM Plex Mono",monospace;font-weight:500;font-size:15px;letter-spacing:.02em}
.cd{font-size:12.5px;color:var(--ink-3)}
.cm{font-size:12.5px;color:var(--ink-2);margin-bottom:8px}
.cn{font-family:"IBM Plex Mono",monospace;font-size:30px;font-weight:500;color:var(--accent);
  font-variant-numeric:tabular-nums;line-height:1.1}
.cl{font-size:11px;letter-spacing:.09em;text-transform:uppercase;color:var(--ink-3)}
.call{background:var(--diff-bg);border:1px solid var(--diff);border-left-width:3px;border-radius:3px;
  padding:20px 22px;display:flex;flex-direction:column;gap:8px}
.call h2{margin:0;font-family:Spectral,Georgia,serif;font-size:20px;font-weight:600;color:var(--diff)}
.call p{margin:0;color:var(--ink);max-width:78ch}
.call code{font-family:"IBM Plex Mono",monospace;font-size:.9em;background:var(--surface);
  padding:1px 5px;border-radius:2px;border:1px solid var(--rule)}
.tw{overflow-x:auto;background:var(--surface);border:1px solid var(--rule);border-radius:3px;
  box-shadow:var(--shadow)}
table{border-collapse:collapse;width:100%;min-width:940px;font-size:13.5px}
thead th{position:sticky;top:0;z-index:2;background:var(--sunk);color:var(--ink-2);
  font-size:11px;letter-spacing:.1em;text-transform:uppercase;font-weight:600;
  text-align:left;padding:12px 14px;border-bottom:1px solid var(--rule-2);white-space:nowrap}
tbody th,tbody td{padding:11px 14px;border-bottom:1px solid var(--rule);vertical-align:top;text-align:left}
tbody td{font-family:"IBM Plex Mono",monospace;font-size:12.5px;color:var(--ink);
  font-variant-numeric:tabular-nums}
tbody th{font-weight:500;min-width:210px;border-left:3px solid transparent}
.par{display:block}
.src{display:block;font-size:11px;line-height:1.4;color:var(--ink-3);font-weight:400;margin-top:3px;
  font-family:"IBM Plex Mono",monospace}
tr.sec th{background:var(--accent-soft);color:var(--accent);font-family:Spectral,Georgia,serif;
  font-size:14px;font-weight:600;letter-spacing:.02em;padding:9px 14px;border-left:3px solid var(--accent)}
.st{width:1%;white-space:nowrap}
.pill{display:inline-block;font-family:"IBM Plex Sans",sans-serif;font-size:10.5px;font-weight:600;
  letter-spacing:.08em;padding:3px 9px;border-radius:2px;border:1px solid currentColor}
.s-same th{border-left-color:var(--ok)} .s-same .pill{color:var(--ok);background:var(--ok-bg)}
.s-pic th{border-left-color:var(--warn)} .s-pic .pill{color:var(--warn);background:var(--warn-bg)}
.s-differs th{border-left-color:var(--diff)} .s-differs .pill{color:var(--diff);background:var(--diff-bg)}
.flag{display:inline-block;margin-top:4px;font-size:10px;font-weight:600;letter-spacing:.08em;
  color:var(--diff);background:var(--diff-bg);border:1px solid var(--diff);border-radius:2px;padding:2px 7px}
.legend{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:14px}
.leg{background:var(--surface);border:1px solid var(--rule);border-radius:3px;padding:14px 16px}
.leg b{display:inline-block;font-size:10.5px;letter-spacing:.08em;padding:3px 9px;border-radius:2px;
  border:1px solid currentColor;margin-bottom:8px}
.leg p{margin:0;font-size:13px;color:var(--ink-2)}
.foot{border-top:1px solid var(--rule);padding-top:20px;color:var(--ink-3);font-size:12.5px;
  display:flex;flex-direction:column;gap:6px}
.foot code{font-family:"IBM Plex Mono",monospace;color:var(--ink-2)}
@media (prefers-reduced-motion:no-preference){.chip,.leg{transition:border-color .15s ease}}
.chip:hover,.leg:hover{border-color:var(--rule-2)}
</style>
<div class="wrap">
<header>
  <p class="eyebrow">LaserProdShock &middot; Phase 4 &middot; cross-code audit</p>
  <h1>What FLASH, PSC and WarpX actually share</h1>
  <p class="lede">Three codes on one laser-ablation problem, parameter by parameter. The campaign spent a dozen runs chasing a 3&times; plume-temperature gap as physics. It is <em>one row</em> &mdash; and the normalised axes hid it, because both PIC codes agree on <em>m<sub>i</sub>/m<sub>e</sub></em> while disagreeing on the ion itself.</p>
  <p class="lede" style="margin-top:14px">The PSC column is <code>run_ourflash_511keV</code> &mdash; <em>m<sub>e</sub>c&sup2;</em> = 511&nbsp;keV, the real electron rest energy, and collisions at exactly 1.00&times; physical. The paper's 60&nbsp;keV leg cuts <em>c</em> a further 2.92&times; beyond the similarity transform and runs 72.5&times; over-collisional, so it is not the leg to isolate WarpX against. The two legs share every length and every physical mass, and agree to 1.4% on plume <em>T<sub>e</sub></em>.</p>
</header>
<section class="chips">@@CHIPS@@</section>
<section class="call">
  <h2>The row that explains it</h2>
  <p>PSC reaches <code>m_i/m_e = 100</code> by making the <strong>electron 18.4&times; heavier</strong>, keeping a <strong>real aluminium ion</strong>. WarpX makes the <strong>ion 18.2&times; lighter</strong>, keeping a real electron. Same ratio, opposite implementation. Only the <em>ion</em> mass sets <code>C_S</code> and the Manheimer steady state &mdash; so the whole temperature gap follows. Measured against its own steady state, every code lands at 0.86&ndash;1.02, and the two PIC codes land on the <em>same</em> 1.02. <strong>None of them is anomalous.</strong></p>
</section>
<div class="tw"><table>
<thead><tr><th>Parameter &amp; provenance</th><th>FLASH</th><th>PSC</th><th>WarpX</th><th>Status</th></tr></thead>
<tbody>@@ROWS@@</tbody></table></div>
<section class="legend">
  <div class="leg"><b style="color:var(--ok);background:var(--ok-bg)">SAME</b><p>Identical across all three by design. If one of these ever differs, that is a bug.</p></div>
  <div class="leg"><b style="color:var(--warn);background:var(--warn-bg)">PIC</b><p>The two PIC codes agree; FLASH differs because it is a fluid code. Expected.</p></div>
  <div class="leg"><b style="color:var(--diff);background:var(--diff-bg)">DIFFERS</b><p>The codes disagree in a way that matters and is not merely a unit convention.</p></div>
</section>
<footer class="foot">
  <div>Generated by <code>scripts/xcode_matrix.py</code> &mdash; the spreadsheet and this page come from the same rows and cannot diverge.</div>
  <div>Spreadsheet: <code>media/xcode/xcode_matrix.xlsx</code> and <code>.csv</code></div>
  <div>Legs: FLASH <code>DELIVERY.md</code> &middot; PSC <code>~/psc-raytrace/run_ourflash_511keV</code> @ 99.66&nbsp;ps &middot; WarpX <code>runs/P4/P4_lez_kin_mr100</code> @ 10.93&nbsp;ps. <code>f_abs</code> is the final instantaneous value in every leg.</div>
  <div>Temperatures are n-weighted over 0.05 &lt; n<sub>e</sub>/n<sub>cr</sub> &lt; 1.0, each leg on its own &zeta;. Steady state is 823&nbsp;eV &times; (m<sub>i</sub>/m<sub>i,real</sub>)<sup>1/3</sup> &times; f<sub>abs</sub><sup>2/3</sup>.</div>
</footer>
</div>
"""

if __name__ == "__main__":
    # HTML is emitted here, after TEMPLATE/write_html are defined.
    write_html(f"{OUT}/xcode_matrix.html")
    print(f"wrote {OUT}/xcode_matrix.html")
